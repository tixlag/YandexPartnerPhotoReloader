import sys
import time
import re
import os
import io
import sqlite3
import pickle
import tempfile
import threading
from dataclasses import dataclass
from typing import List, Optional, Dict, Tuple

import pandas as pd
import requests
from openpyxl import load_workbook
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtCore import Qt
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from selenium.webdriver.common.action_chains import ActionChains

# ======================
# Persistent storage
# ======================
DB_FILE = "processed_items.sqlite3"
COOKIES_FILE = "cookies.pkl"
DEFAULT_WAIT = 20
RETRY_COUNT = 3
RETRY_DELAY = 5

SKU_COLUMN_CANDIDATES = [
    "Ваш SKU *", "Ваш SKU", "SKU", "Артикул", "Артикул продавца", "Article"
]

@dataclass
class Cabinet:
    business_id: str 
    name: str
    dashboard_href: str  # /business/{id}/dashboard?view=marketplace

class Storage:
    def __init__(self, path: str = DB_FILE):
        self.conn = sqlite3.connect(path, check_same_thread=False)
        self._init_schema()
        self.lock = threading.Lock()

    def _init_schema(self):
        cur = self.conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS processed (
                campaign_id TEXT NOT NULL,
                sku TEXT NOT NULL,
                processed_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (campaign_id, sku)
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS cabinets (
                business_id TEXT PRIMARY KEY,
                name TEXT,
                dashboard_href TEXT
            )
            """
        )
        self.conn.commit()

    def add_processed(self, campaign_id: str, sku: str):
        with self.lock:
            self.conn.execute(
                "INSERT OR IGNORE INTO processed (campaign_id, sku) VALUES (?, ?)",
                (campaign_id, sku),
            )
            self.conn.commit()

    def is_processed(self, campaign_id: str, sku: str) -> bool:
        with self.lock:
            row = self.conn.execute(
                "SELECT 1 FROM processed WHERE campaign_id=? AND sku=?",
                (campaign_id, sku),
            ).fetchone()
            return row is not None

    def bulk_mark_processed(self, campaign_id: str, skus: List[str]):
        with self.lock:
            self.conn.executemany(
                "INSERT OR IGNORE INTO processed (campaign_id, sku) VALUES (?, ?)",
                [(campaign_id, s) for s in skus],
            )
            self.conn.commit()

    def save_cabinets(self, cabinets: List[Cabinet]):
        with self.lock:
            self.conn.executemany(
                "INSERT OR REPLACE INTO cabinets (business_id, name, dashboard_href) VALUES (?, ?, ?)",
                [(c.business_id, c.name, c.dashboard_href) for c in cabinets],
            )
            self.conn.commit()

    def load_cabinets(self) -> List[Cabinet]:
        with self.lock:
            rows = self.conn.execute(
                "SELECT business_id, name, dashboard_href FROM cabinets"
            ).fetchall()
        return [Cabinet(*r) for r in rows]


# ======================
# Selenium driver wrapper
# ======================
class YandexMarketPhotoReuploadeDriver:
    def __init__(self, log_fn):
        self.driver = None
        self.actions = None
        self.wait: Optional[WebDriverWait] = None
        self.log = log_fn

    def start(self):
        chrome_options = Options()
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        # chrome_options.add_argument("--headless=new")  # GUI is needed for captcha; leave commented

        self.driver = webdriver.Chrome(options=chrome_options)
        self.actions = ActionChains(self.driver)
        self.wait = WebDriverWait(self.driver, DEFAULT_WAIT)
        self.log("Браузер запущен")

    def stop(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None
            self.log("Браузер остановлен")

    # -------------
    # Auth & Cookies
    # -------------
    def open_home(self):
        self.driver.get("https://partner.market.yandex.ru/")

    def save_cookies(self):
        try:
            cookies = self.driver.get_cookies()
            with open(COOKIES_FILE, "wb") as f:
                pickle.dump(cookies, f)
            self.log("Cookies сохранены")
        except Exception as e:
            self.log(f"Не удалось сохранить cookies: {e}")

    def load_cookies(self):
        if not os.path.exists(COOKIES_FILE):
            self.log("Файл cookies не найден — авторизуйтесь вручную")
            return False
        try:
            self.driver.get("https://partner.market.yandex.ru/")
            with open(COOKIES_FILE, "rb") as f:
                cookies = pickle.load(f)
            for c in cookies:
                # Selenium may require domain without leading dot
                c = c.copy()
                c.pop("sameSite", None)  # some drivers dislike this flag
                try:
                    self.driver.add_cookie(c)
                except Exception:
                    pass
            self.driver.refresh()
            self.log("Cookies загружены")
            return True
        except Exception as e:
            self.log(f"Ошибка загрузки cookies: {e}")
            return False

    # -------------
    # Helpers
    # -------------
    def _is_captcha(self) -> bool:
        try:
            html = self.driver.page_source.lower()
            if "captcha" in html:
                return True
        except Exception:
            return False
        return False

    def ensure_no_captcha(self):
        if self._is_captcha():
            raise RuntimeError("Обнаружена капча. Пожалуйста, решите её в открытом браузере и нажмите Продолжить.")

    @staticmethod
    def extract_ints(text: str) -> Optional[str]:
        m = re.search(r"(\d+)", text or "")
        return m.group(1) if m else None

    # -------------
    # Cabinets scraping from settings page
    # -------------
    def open_business_settings(self, business_id: str):
        url = f"https://partner.market.yandex.ru/business/{business_id}/settings?activeTab=all"
        self.driver.get(url)
        self.wait.until(lambda d: "settings" in d.current_url)
        self.ensure_no_captcha()
        self.log(f"Открыта страница настроек бизнеса {business_id}")

    def scrape_cabinets_from_current_page(self) -> List[Cabinet]:
        try:
            self.driver.get("https://partner.market.yandex.ru/main-redirect")
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div[data-e2e='business-list']"))
            )
        except TimeoutException:
            tabs = self.driver.window_handles
            self.driver.switch_to.window(tabs[-1])
            self.ensure_no_captcha()
            self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div[data-e2e='business-list']")))

        cards = self.driver.find_elements(By.CSS_SELECTOR, "div[data-e2e='business-card-wrapper']")
        cabinets: List[Cabinet] = []

        for card in cards:
            text = card.text
            business_id = None
            name = None

            # --- ID ---
            try:
                bid_el = card.find_element(By.CSS_SELECTOR, "[data-e2e='business-id']")
                business_id = bid_el.text.strip()
            except Exception:
                match = re.search(r"\bID(?:\s+\S+)*\s+(\d+)\b", text, re.IGNORECASE)
                if match:
                    business_id = match.group(1)

            # --- Name ---
            try:
                name_el = card.find_element(By.CSS_SELECTOR, "span[data-e2e='business-card-name']")
                name = name_el.text.strip()
            except Exception:
                text_lines = [line.strip() for line in text.splitlines() if line.strip()]
                if text_lines:
                    # если первая строка содержит ID — убираем его
                    if business_id and business_id in text_lines[0]:
                        name = " ".join([w for w in text_lines[0].split() if business_id not in w])
                    else:
                        name = text_lines[0]
                else:
                    name = f"Бизнес {business_id or '?'}"

            # --- Ссылка в кабинет ---
            dashboard_link = ""
            if business_id:
                try:
                    link_el = self.driver.find_element(
                        By.XPATH, f"//a[contains(@href,'/business/{business_id}/dashboard')]"
                    )
                    dashboard_link = link_el.get_attribute("href")
                except Exception:
                    try:
                        link_el2 = card.find_element(
                            By.XPATH, ".//following::a[contains(@href,'/business/') and contains(@href,'/dashboard')][1]"
                        )
                        dashboard_link = link_el2.get_attribute("href")
                    except Exception:
                        dashboard_link = ""

            cabinets.append(Cabinet(business_id=business_id, name=name, dashboard_href=dashboard_link))

        if not cabinets:
            self.log("Не найдено ни одного кабинета на странице — проверьте, верная ли страница.")
        else:
            self.log(f"Найдено кабинетов: {len(cabinets)}")

        return cabinets


    def get_campaign_id_from_business(self, business_id: str) -> Optional[str]:
        # Откроем страницу, где есть ссылка /business/{id}/showcase?campaignId=...
        url = f"https://partner.market.yandex.ru/business/{business_id}"
        self.driver.get(url)
        try:
            self.wait.until(EC.presence_of_element_located((By.XPATH, f"//a[contains(@href,'/business/{business_id}/showcase') and contains(@href,'campaignId=')]")))
        except TimeoutException:
            # иногда ссылка доступна из меню/переключения — попробуем перейти в разделы
            pass

        self.ensure_no_captcha()
        links = self.driver.find_elements(By.XPATH, f"//a[contains(@href,'/business/{business_id}/showcase') and contains(@href,'campaignId=')]")
        if not links:
            # Возможно, ссылка доступна из меню. Попробуем искать глобально
            links = self.driver.find_elements(By.XPATH, "//a[contains(@href,'/showcase?campaignId=')]")
        for a in links:
            href = a.get_attribute("href")
            m = re.search(r"campaignId=(\d+)", href or "")
            if m:
                cid = m.group(1)
                self.log(f"Найден campaignId: {cid}")
                return cid
        self.log("Не удалось найти campaignId для выбранного кабинета")
        return None

    # -------------
    # Offer page -> open picture modal -> download last image -> upload back -> save
    # -------------
    def _requests_session_with_cookies(self) -> requests.Session:
        s = requests.Session()
        for c in self.driver.get_cookies():
            try:
                s.cookies.set(c.get('name'), c.get('value'), domain=c.get('domain'), path=c.get('path'))
            except Exception:
                # fallback without domain/path
                s.cookies.set(c.get('name'), c.get('value'))
        return s

    def _download_image(self, url: str, dest_path: str):
        # url может быть //avatars.mds.yandex.net/...
        if url.startswith("//"):
            url = "https:" + url
        sess = self._requests_session_with_cookies()
        r = sess.get(url, timeout=60)
        r.raise_for_status()
        with open(dest_path, "wb") as f:
            f.write(r.content)

    def process_sku(self, campaign_id: str, sku: str, logger) -> bool:
        self.ensure_no_captcha()
        offer_url = f"https://partner.market.yandex.ru/supplier/{campaign_id}/assortment/offer-card?article={requests.utils.quote(sku)}&source=businessStocks"
        self.driver.get(offer_url)

        try:
            # Ждем превью картинок
            self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "img.styles-picture___6gWHl")))
        except TimeoutException:
            logger(f"[{sku}] Нет превью изображений на карточке")
            self.ensure_no_captcha()
            return False

        self.ensure_no_captcha()

        attempts = 0
        success = False
        while attempts < RETRY_COUNT and not success:
            attempts += 1
            # Клик по первому превью
            try:
                first_thumb_fake = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "img.styles-picture___6gWHl")))
                first_thumb_real = self.driver.find_elements(By.CSS_SELECTOR, "span.styles-layout___1YRjC")[0]
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", first_thumb_fake)
                
                # Навести мышь на элемент
                self.actions.move_to_element(first_thumb_fake).click().perform()
                # first_thumb_real.click()
            except Exception as e:
                logger(f"[{sku}] Не удалось кликнуть превью: {e}")
                self.ensure_no_captcha()

            # Ждем модалку
            try:
                modal = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.___wrapper___7pLKs.style-picturesDrawer___55UwA")))
                success = True
            except TimeoutException:
                logger(f"[{sku}] Модальное окно с картинками не открылось")
                self.ensure_no_captcha()
            if not success:
                 logger(f"[{sku}] Модальное окно с картинками не открылось")
                 return False

        # Найти все большие изображения в модалке, взять последнее
        try:
            self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "img[data-testid='loaded-image']")))
            big_images = modal.find_elements(By.CSS_SELECTOR, "img.style-root___17qgj.style-main___6BATS")
            if not big_images:
                logger(f"[{sku}] Не найдено больших изображений в модалке")
                return False
            last_img = big_images[-1]
            src = last_img.get_attribute("src")
            if not src:
                logger(f"[{sku}] Не удалось получить src последнего изображения")
                return False
        except Exception as e:
            logger(f"[{sku}] Ошибка получения изображения: {e}")
            return False

        # Скачиваем во временный файл
        tmp_dir = tempfile.mkdtemp(prefix="ym_images_")
        tmp_path = os.path.join(tmp_dir, "orig.webp")
        try:
            self._download_image(src, tmp_path)
            logger(f"[{sku}] Изображение скачано: {tmp_path}")
        except Exception as e:
            logger(f"[{sku}] Не удалось скачать изображение: {e}")
            return False

        # Нажимаем на первый элемент <span class="___content___2ml2l"> в модалке
        try:
            upload_trigger = self.driver.find_elements(By.CSS_SELECTOR, "span.___content___2ml2l")[0]
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", upload_trigger)
            # upload_trigger.click()
        except Exception as e:
            logger(f"[{sku}] Не удалось кликнуть кнопку загрузки: {e}")
            return False

        # Ищем input[type=file] и отправляем файл
        try:
            # file_input = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']:nth-of-type(2)")))
           
            file_inputs = self.wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "input[type='file']")))
            if len(file_inputs) >= 2:
                file_input = file_inputs[1]  # второй элемент (индекс 1)
            else:
                raise Exception("Недостаточно file input элементов")

            file_input.send_keys(tmp_path)
            logger(f"[{sku}] Файл отправлен на загрузку")
        except TimeoutException:
            logger(f"[{sku}] Не найден input[type=file] для загрузки")
            return False
        except Exception as e:
            logger(f"[{sku}] Ошибка отправки файла: {e}")
            return False

        # Подождать завершения загрузки — эвристика: ждём исчезновение индикаторов прогресса или появление превью
        time.sleep(2)
        self.ensure_no_captcha()

        # Закрыть модалку
        try:
            close_btn = self.driver.find_element(By.XPATH, "//span[@aria-label='Закрыть']")
            self.driver.execute_script("arguments[0].click();", close_btn)
        except Exception:
            # не критично — попробуем дальше
            pass

        # Сохранить изменения
        try:
            save_btn = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[data-e2e='next-step-button']")))
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", save_btn)
            save_btn.click()
        except TimeoutException:
            logger(f"[{sku}] Кнопка Сохранить не найдена")
            return False
        except Exception as e:
            logger(f"[{sku}] Ошибка клика по Сохранить: {e}")
            return False

        # Подождать подтверждения — эвристика: дождёмся некоего уведомления или просто пауза
        time.sleep(2)
        self.ensure_no_captcha()
        logger(f"[{sku}] Сохранение инициировано (ожидаем успешный ответ)")
        return True


# ======================
# Worker thread for long tasks
# ======================
class ProcessWorker(QtCore.QThread):
    log_signal = QtCore.pyqtSignal(str)
    progress_signal = QtCore.pyqtSignal(int, int)
    finished_signal = QtCore.pyqtSignal()
    captcha_signal = QtCore.pyqtSignal(str)

    def __init__(self, driver: YandexMarketPhotoReuploadeDriver, storage: Storage, campaign_id: str, skus: List[str], skip_processed: bool):
        super().__init__()
        self.driver = driver
        self.storage = storage
        self.campaign_id = campaign_id
        self.skus = skus
        self.skip_processed = skip_processed
        self._pause_for_captcha = False
        self._abort = False

    def run(self):
        total = len(self.skus)
        for idx, sku in enumerate(self.skus, start=1):
            if self._abort:
                break

            # Пропуск уже обработанных
            if self.skip_processed and self.storage.is_processed(self.campaign_id, sku):
                self.log_signal.emit(f"Пропуск {sku} — уже обработан ранее")
                self.progress_signal.emit(idx, total)
                continue

            attempts = 0
            success = False
            while attempts < RETRY_COUNT and not success and not self._abort:
                attempts += 1
                try:
                    # Проверка капчи заранее
                    if self.driver._is_captcha():
                        self.captcha_signal.emit("Обнаружена капча. Решите её в открытом браузере и нажмите 'Продолжить'.")
                        # Ожидание разблокировки
                        while self._pause_for_captcha and not self._abort:
                            time.sleep(0.5)

                    self.log_signal.emit(f"[{sku}] Попытка {attempts}/{RETRY_COUNT}")
                    success = self.driver.process_sku(self.campaign_id, sku, self.log_signal.emit)
                    if not success:
                        time.sleep(RETRY_DELAY)
                except RuntimeError as e:
                    # капча
                    self.captcha_signal.emit(str(e))
                    while self._pause_for_captcha and not self._abort:
                        time.sleep(0.5)
                except Exception as e:
                    self.log_signal.emit(f"[{sku}] Ошибка: {e}")
                    time.sleep(RETRY_DELAY)

            if success:
                self.storage.add_processed(self.campaign_id, sku)
                self.log_signal.emit(f"[{sku}] УСПЕХ — отмечен как обработанный")
            else:
                self.log_signal.emit(f"[{sku}] НЕ УДАЛОСЬ обработать после {RETRY_COUNT} попыток")

            self.progress_signal.emit(idx, total)

        self.finished_signal.emit()

    def pause_for_captcha(self):
        self._pause_for_captcha = True

    def resume_after_captcha(self):
        self._pause_for_captcha = False

    def abort(self):
        self._abort = True


# ======================
# GUI
# ======================
class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Yandex Market Photo Reuploade")
        self.resize(980, 720)

        self.storage = Storage()
        self.driver = YandexMarketPhotoReuploadeDriver(self.log)
        self.worker: Optional[ProcessWorker] = None
        self.campaign_id: Optional[str] = None
        self.current_business_id: Optional[str] = None

        self._build_ui()

    # ---------- UI ----------
    def _build_ui(self):
        central = QtWidgets.QWidget()
        self.setCentralWidget(central)

        layout = QtWidgets.QVBoxLayout(central)

        # Top controls
        top_bar = QtWidgets.QHBoxLayout()

        self.btn_start_browser = QtWidgets.QPushButton("Запустить браузер")
        self.btn_start_browser.clicked.connect(self.on_start_browser)
        top_bar.addWidget(self.btn_start_browser)

        self.btn_open_home = QtWidgets.QPushButton("Открыть главную")
        self.btn_open_home.clicked.connect(self.on_open_home)
        self.btn_open_home.setEnabled(False)
        top_bar.addWidget(self.btn_open_home)

        self.btn_save_cookies = QtWidgets.QPushButton("Сохранить cookies")
        self.btn_save_cookies.clicked.connect(self.on_save_cookies)
        self.btn_save_cookies.setEnabled(False)
        top_bar.addWidget(self.btn_save_cookies)

        self.btn_load_cookies = QtWidgets.QPushButton("Загрузить cookies")
        self.btn_load_cookies.clicked.connect(self.on_load_cookies)
        self.btn_load_cookies.setEnabled(False)
        top_bar.addWidget(self.btn_load_cookies)

        layout.addLayout(top_bar)

        # Business controls
        business_box = QtWidgets.QGroupBox("Кабинеты")
        business_layout = QtWidgets.QGridLayout(business_box)

        # self.edit_business_id = QtWidgets.QLineEdit()
        # self.edit_business_id.setPlaceholderText("ID кабинета (business_id) для открытия страницы настроек")
        # self.btn_open_settings = QtWidgets.QPushButton("Открыть страницу настроек")
        # self.btn_open_settings.clicked.connect(self.on_open_settings)
        # self.btn_open_settings.setEnabled(False)

        # self.btn_scan_cabinets = QtWidgets.QPushButton("Считать кабинеты с текущей страницы")
        # self.btn_scan_cabinets.clicked.connect(self.on_scan_cabinets)
        # self.btn_scan_cabinets.setEnabled(False)

        self.combo_cabinets = QtWidgets.QComboBox()
        self.btn_pick_campaign = QtWidgets.QPushButton("Получить campaignId для выбранного кабинета")
        self.btn_pick_campaign.clicked.connect(self.on_pick_campaign)
        self.btn_pick_campaign.setEnabled(False)

        # business_layout.addWidget(QtWidgets.QLabel("Business ID:"), 0, 0)
        # business_layout.addWidget(self.edit_business_id, 0, 1)
        # business_layout.addWidget(self.btn_open_settings, 0, 2)
        # business_layout.addWidget(self.btn_scan_cabinets, 1, 0, 1, 3)
        business_layout.addWidget(QtWidgets.QLabel("Найденные кабинеты:"), 2, 0)
        business_layout.addWidget(self.combo_cabinets, 2, 1)
        business_layout.addWidget(self.btn_pick_campaign, 2, 2)

        layout.addWidget(business_box)

        # File & process controls
        proc_box = QtWidgets.QGroupBox("Обработка товаров")
        proc_layout = QtWidgets.QGridLayout(proc_box)

        self.btn_load_xlsx = QtWidgets.QPushButton("Загрузить XLSX (остатки)")
        self.btn_load_xlsx.clicked.connect(self.on_load_xlsx)
        self.btn_load_xlsx.setEnabled(False)

        self.chk_skip_processed = QtWidgets.QCheckBox("Пропускать уже обработанные")
        self.chk_skip_processed.setChecked(True)

        self.list_skus = QtWidgets.QListWidget()

        self.btn_start = QtWidgets.QPushButton("Запустить обработку")
        self.btn_start.clicked.connect(self.on_start_processing)
        self.btn_start.setEnabled(False)

        self.btn_abort = QtWidgets.QPushButton("Прервать")
        self.btn_abort.clicked.connect(self.on_abort)
        self.btn_abort.setEnabled(False)

        self.btn_continue_after_captcha = QtWidgets.QPushButton("Продолжить после капчи")
        self.btn_continue_after_captcha.clicked.connect(self.on_continue_after_captcha)
        self.btn_continue_after_captcha.setEnabled(False)

        self.progress = QtWidgets.QProgressBar()

        proc_layout.addWidget(self.btn_load_xlsx, 0, 0)
        proc_layout.addWidget(self.chk_skip_processed, 0, 1)
        proc_layout.addWidget(self.list_skus, 1, 0, 1, 3)
        proc_layout.addWidget(self.btn_start, 2, 0)
        proc_layout.addWidget(self.btn_abort, 2, 1)
        proc_layout.addWidget(self.btn_continue_after_captcha, 2, 2)
        proc_layout.addWidget(self.progress, 3, 0, 1, 3)

        layout.addWidget(proc_box)

        # Log
        self.log_view = QtWidgets.QTextEdit()
        self.log_view.setReadOnly(True)
        layout.addWidget(self.log_view, stretch=1)

        # preload saved cabinets
        self._reload_cabinets_combo()

    # ---------- Buttons handlers ----------
    def on_start_browser(self):
        try:
            self.driver.start()
            self.btn_open_home.setEnabled(True)
            self.btn_save_cookies.setEnabled(True)
            self.btn_load_cookies.setEnabled(True)
            # self.btn_open_settings.setEnabled(True)
            # self.btn_scan_cabinets.setEnabled(True)
            self.btn_pick_campaign.setEnabled(True)
            self.btn_load_xlsx.setEnabled(True)
            self.driver.load_cookies()
            time.sleep(2)
            self.on_scan_cabinets()
        except WebDriverException as e:
            self.log(f"Не удалось запустить браузер: {e}")

    def on_open_home(self):
        self.driver.open_home()

    def on_save_cookies(self):
        self.driver.save_cookies()
        self.on_scan_cabinets()

    def on_load_cookies(self):
        ok = self.driver.load_cookies()
        if not ok:
            self.log("Cookies не загружены — выполните вход вручную и сохраните их")

    def on_open_settings(self):
        bid = self.edit_business_id.text().strip()
        if not bid:
            self.log("Укажите business_id")
            return
        try:
            self.driver.open_business_settings(bid)
            self.current_business_id = bid
        except Exception as e:
            self.log(str(e))

    def on_scan_cabinets(self):
        try:
            cabins = self.driver.scrape_cabinets_from_current_page()
            if cabins:
                self.storage.save_cabinets(cabins)
                self._reload_cabinets_combo(cabins)
        except Exception as e:
            self.log(str(e))

    def _reload_cabinets_combo(self, cabins: Optional[List[Cabinet]] = None):
        if cabins is None:
            cabins = self.storage.load_cabinets()
        self.combo_cabinets.clear()
        for c in cabins:
            self.combo_cabinets.addItem(f"{c.name} (ID {c.business_id})", userData=c)

    def on_pick_campaign(self):
        data = self.combo_cabinets.currentData()
        if not data:
            self.log("Сначала считайте и выберите кабинет")
            return
        self.current_business_id = data.business_id
        cid = self.driver.get_campaign_id_from_business(self.current_business_id)
        if cid:
            self.campaign_id = cid
            self.log(f"Текущий campaignId: {self.campaign_id}")
            self.btn_start.setEnabled(True)
        else:
            self.log("Не удалось определить campaignId")

    def on_load_xlsx(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Выберите XLSX", "", "Excel (*.xlsx)")
        if not path:
            return

        # определяем выбранный бизнес (если есть)
        selected_business_id = None
        if self.current_business_id:
            selected_business_id = str(self.current_business_id)
        else:
            try:
                data = self.combo_cabinets.currentData()
                if data:
                    selected_business_id = str(getattr(data, "business_id", "") or "")
            except Exception:
                selected_business_id = None

        reported_business_id = None
        header_row_idx = None

        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb.active

            # B3 — ID кабинета в файле (если есть)
            try:
                b3 = ws["B3"].value
                if b3 is not None:
                    m = re.search(r"(\d+)", str(b3))
                    if m:
                        reported_business_id = m.group(1)
            except Exception:
                reported_business_id = None

            # Попробуем найти строку заголовков динамически (ищем ячейку с "Ваш SKU")
            for r in range(1, 30):  # первые ~30 строк
                row_vals = [str(c.value).strip() if c.value is not None else '' for c in ws[r]]
                if any('ваш sku' in v.lower() for v in row_vals):
                    header_row_idx = r
                    break

            # Если не нашли — по умолчанию 7-я строка
            if header_row_idx is None:
                header_row_idx = 7

        except Exception as e:
            self.log(f"Не удалось прочитать шапку XLSX через openpyxl: {e}")
            header_row_idx = 7
        finally:
            try:
                wb.close()
            except Exception:
                pass

        # сверка ID кабинета
        if selected_business_id and reported_business_id and selected_business_id != reported_business_id:
            reply = QtWidgets.QMessageBox.question(
                self,
                "Несовпадение ID кабинета",
                f"В файле указан ID кабинета: {reported_business_id},\nа выбран кабинет: {selected_business_id}.\n\nПродолжить загрузку?",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No,
            )
            if reply == QtWidgets.QMessageBox.No:
                self.log("Загрузка XLSX отменена пользователем из-за несовпадения ID кабинета.")
                return
        elif selected_business_id and not reported_business_id:
            self.log("Внимание: не удалось прочитать ID кабинета из ячейки B3. Продолжаем без проверки.")

        # Читаем файл через openpyxl, так как pandas плохо работает с объединенными ячейками
        try:
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            
            # Функция для получения значения из объединенной ячейки
            def get_merged_cell_value(row, col):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    return str(cell.value).strip()
                
                # Поиск в объединенных диапазонах
                for merged_range in ws.merged_cells.ranges:
                    if merged_range.min_row <= row <= merged_range.max_row and \
                    merged_range.min_col <= col <= merged_range.max_col:
                        # Берем значение из верхней левой ячейки объединенного диапазона
                        top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        return str(top_left_cell.value).strip() if top_left_cell.value is not None else ""
                return ""

            # Определяем максимальную колонку с данными
            max_col = 0
            for row in ws.iter_rows(min_row=header_row_idx, max_row=min(header_row_idx + 50, ws.max_row)):
                for i, cell in enumerate(row, 1):
                    if cell.value is not None and str(cell.value).strip():
                        max_col = max(max_col, i)
            
            if max_col == 0:
                max_col = ws.max_column or 20

            # Собираем заголовки
            headers = []
            for col in range(1, max_col + 1):
                header_val = get_merged_cell_value(header_row_idx, col)
                if not header_val:
                    # Пробуем соседние строки
                    for offset in [-1, 1, 2]:
                        alt_row = header_row_idx + offset
                        if alt_row >= 1:
                            alt_val = get_merged_cell_value(alt_row, col)
                            if alt_val:
                                header_val = alt_val
                                break
                
                if not header_val:
                    header_val = f"Column_{col}"
                headers.append(header_val)

            # Собираем данные
            data_rows = []
            start_data_row = header_row_idx + 1
            
            for row_num in range(start_data_row, ws.max_row + 1):
                row_data = []
                has_data = False
                
                for col in range(1, len(headers) + 1):
                    cell_val = get_merged_cell_value(row_num, col)
                    if cell_val and cell_val.lower() != "nan":
                        has_data = True
                    row_data.append(cell_val if cell_val else "")
                
                if has_data:
                    data_rows.append(row_data)
                elif len(data_rows) > 0:
                    # Если встретили пустую строку после того, как уже были данные,
                    # проверим еще несколько строк, чтобы убедиться что данные закончились
                    empty_count = 0
                    for check_row in range(row_num, min(row_num + 5, ws.max_row + 1)):
                        check_has_data = False
                        for check_col in range(1, len(headers) + 1):
                            check_val = get_merged_cell_value(check_row, check_col)
                            if check_val and check_val.lower() != "nan":
                                check_has_data = True
                                break
                        if not check_has_data:
                            empty_count += 1
                        else:
                            break
                    
                    if empty_count >= 5:  # Если 5 пустых строк подряд - прекращаем
                        break

            # Создаем DataFrame
            if data_rows:
                # Обрезаем заголовки до реального количества колонок в данных
                max_data_cols = max(len(row) for row in data_rows) if data_rows else 0
                if max_data_cols < len(headers):
                    headers = headers[:max_data_cols]
                elif max_data_cols > len(headers):
                    # Добавляем недостающие заголовки
                    for i in range(len(headers), max_data_cols):
                        headers.append(f"Column_{i + 1}")
                
                # Приводим все строки к одинаковой длине
                for i, row in enumerate(data_rows):
                    while len(row) < len(headers):
                        row.append("")
                    data_rows[i] = row[:len(headers)]
                
                df = pd.DataFrame(data_rows, columns=headers)
            else:
                df = pd.DataFrame()

        except Exception as e:
            self.log(f"Ошибка при чтении файла через openpyxl: {e}")
            # Fallback к pandas
            try:
                df = pd.read_excel(path, sheet_name=0, skiprows=header_row_idx - 1, engine="openpyxl")
            except Exception as e2:
                self.log(f"Не удалось прочитать файл Excel: {e2}")
                return
        finally:
            try:
                wb.close()
            except Exception:
                pass

        # Очистка пустых строк/столбцов
        if not df.empty:
            df = df.dropna(how="all").dropna(axis=1, how="all")
        
        if df.empty:
            self.log("После очистки данных таблица пуста.")
            return

        # Нормализация имён колонок и поиск колонки SKU
        def norm(s: str) -> str:
            return re.sub(r"[^a-zа-я0-9]+", " ", str(s).strip().lower())

        sku_col = None
        for col in df.columns:
            nc = norm(col)
            if nc.startswith("ваш sku") or nc == "sku" or nc == "артикул":
                sku_col = col
                break

        if sku_col is None:
            # запасной поиск
            SKU_COLUMN_CANDIDATES = ["Ваш SKU", "SKU", "Артикул", "Код товара"]
            for cand in SKU_COLUMN_CANDIDATES:
                cand_n = norm(cand)
                for col in df.columns:
                    if norm(col).startswith(cand_n):
                        sku_col = col
                        break
                if sku_col:
                    break

        if sku_col is None:
            self.log("Не найдена колонка со SKU. Убедитесь, что столбец называется 'Ваш SKU *' или аналогично.")
            self.log(f"Найденные колонки: {list(df.columns)}")
            return

        # Список SKU
        series = df[sku_col]
        skus = [str(x).strip() for x in series.dropna().astype(str).tolist() 
                if str(x).strip() and str(x).strip().lower() != "nan"]

        # Вывод в UI
        self.list_skus.clear()
        for s in skus:
            self.list_skus.addItem(s)

        msg = f"Загружено SKU: {len(skus)}"
        if reported_business_id:
            msg += f" (ID из файла: {reported_business_id})"
        if selected_business_id:
            msg += f"; выбран кабинет: {selected_business_id}"
        self.log(msg)


    def on_start_processing(self):
            if not self.campaign_id:
                self.log("Сначала получите campaignId для кабинета")
                return
            skus = [self.list_skus.item(i).text() for i in range(self.list_skus.count())]
            if not skus:
                self.log("Список SKU пуст")
                return

            self.worker = ProcessWorker(
                driver=self.driver,
                storage=self.storage,
                campaign_id=self.campaign_id,
                skus=skus,
                skip_processed=self.chk_skip_processed.isChecked(),
            )
            self.worker.log_signal.connect(self.log)
            self.worker.progress_signal.connect(self.on_progress)
            self.worker.finished_signal.connect(self.on_finished)
            self.worker.captcha_signal.connect(self.on_captcha)
            self.worker.start()

            self.btn_start.setEnabled(False)
            self.btn_abort.setEnabled(True)
            self.btn_continue_after_captcha.setEnabled(True)

    def on_abort(self):
        if self.worker:
            self.worker.abort()
            self.log("Запрошено прерывание")

    def on_continue_after_captcha(self):
        if self.worker:
            self.worker.resume_after_captcha()
            self.log("Продолжаем после капчи")

    def on_progress(self, done: int, total: int):
        self.progress.setMaximum(total)
        self.progress.setValue(done)

    def on_finished(self):
        self.log("Обработка завершена")
        self.btn_abort.setEnabled(False)
        self.btn_start.setEnabled(True)

    def on_captcha(self, msg: str):
        self.log(msg)
        if self.worker:
            self.worker.pause_for_captcha()

    # ---------- Utils ----------
    def log(self, text: str):
        timestamp = time.strftime("%H:%M:%S")
        self.log_view.append(f"[{timestamp}] {text}")
        self.log_view.moveCursor(QtGui.QTextCursor.End)

    def closeEvent(self, event: QtGui.QCloseEvent):
        try:
            if self.worker and self.worker.isRunning():
                self.worker.abort()
                self.worker.wait(2000)
        except Exception:
            pass
        try:
            self.driver.stop()
        except Exception:
            pass
        event.accept()


def main():
    app = QtWidgets.QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
